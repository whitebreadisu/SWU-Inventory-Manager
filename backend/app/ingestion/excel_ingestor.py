"""Excel inventory ingestion pipeline for SWU Inventory Manager (F4).

Reads the personal collection Excel workbook and populates the inventory table.
Each sheet whose name matches a known set code (SOR, SHD, TWI, JTL, LOF, SEC, LAW)
is processed. Non-set sheets (e.g., ImageData) are skipped.

Idempotent: re-running overwrites existing inventory quantities with values
from the Excel file. Failed card lookups are logged and collected — the import
never aborts on a lookup failure.

Run inside the backend container after CSV ingestion has completed:
    docker compose exec backend python -m app.ingestion.run_inventory_ingestion
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from app.ingestion.normalize import parse_card_number
from app.models.card import Card
from app.models.inventory import Inventory
from app.models.set_model import CardSet

logger = logging.getLogger(__name__)

_KNOWN_SET_CODES = frozenset({"SOR", "SHD", "TWI", "JTL", "LOF", "SEC", "LAW"})

# Maps lowercased Excel column header → variant flag dict.
# Case-insensitive matching is applied in identify_inventory_columns().
# "standard" (LAW) maps identically to "non-foil".
VARIANT_COLUMN_FLAGS: dict[str, dict[str, bool]] = {
    "non-foil": dict(
        is_foil=False,
        is_hyperspace=False,
        is_prestige=False,
        is_showcase=False,
        is_organized_play=False,
    ),
    "standard": dict(
        is_foil=False,
        is_hyperspace=False,
        is_prestige=False,
        is_showcase=False,
        is_organized_play=False,
    ),
    "foil": dict(
        is_foil=True,
        is_hyperspace=False,
        is_prestige=False,
        is_showcase=False,
        is_organized_play=False,
    ),
    "hyperspace": dict(
        is_foil=False,
        is_hyperspace=True,
        is_prestige=False,
        is_showcase=False,
        is_organized_play=False,
    ),
    "f-hyperspace": dict(
        is_foil=True,
        is_hyperspace=True,
        is_prestige=False,
        is_showcase=False,
        is_organized_play=False,
    ),
    "prestige": dict(
        is_foil=False,
        is_hyperspace=False,
        is_prestige=True,
        is_showcase=False,
        is_organized_play=False,
    ),
    "prestige foil": dict(
        is_foil=True,
        is_hyperspace=False,
        is_prestige=True,
        is_showcase=False,
        is_organized_play=False,
    ),
    "promo": dict(
        is_foil=False,
        is_hyperspace=False,
        is_prestige=False,
        is_showcase=False,
        is_organized_play=True,
    ),
    "promo foil": dict(
        is_foil=True,
        is_hyperspace=False,
        is_prestige=False,
        is_showcase=False,
        is_organized_play=True,
    ),
}


@dataclass
class ExcelIngestionResult:
    sheets_processed: int = 0
    inventory_upserted: int = 0
    rows_skipped: int = 0
    lookup_failures: list[dict] = field(default_factory=list)
    sheet_summaries: list[dict] = field(default_factory=list)


def run_excel_ingestion(db: Session, excel_path: Path) -> ExcelIngestionResult:
    """Open the workbook and process each sheet that matches a known set code."""
    result = ExcelIngestionResult()

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    set_map: dict[str, CardSet] = {s.code: s for s in db.query(CardSet).all()}

    for sheet_name in wb.sheetnames:
        if sheet_name not in _KNOWN_SET_CODES:
            logger.debug("Skipping sheet %r", sheet_name)
            continue

        card_set = set_map.get(sheet_name)
        if card_set is None:
            logger.warning(
                "Sheet %r: no matching set in database — skipping", sheet_name
            )
            continue

        ws = wb[sheet_name]
        upserted, skipped, failures = _process_sheet(db, ws, sheet_name, card_set.id)
        db.commit()

        result.sheets_processed += 1
        result.inventory_upserted += upserted
        result.rows_skipped += skipped
        result.lookup_failures.extend(failures)
        result.sheet_summaries.append(
            {
                "set_code": sheet_name,
                "upserted": upserted,
                "skipped": skipped,
                "failed_lookups": len(failures),
            }
        )
        logger.info(
            "Sheet %s: %d upserted, %d skipped, %d lookup failures",
            sheet_name,
            upserted,
            skipped,
            len(failures),
        )

    return result


def _process_sheet(
    db: Session, ws, set_code: str, set_id: int
) -> tuple[int, int, list[dict]]:
    rows = list(ws.iter_rows(values_only=True))

    header_idx = find_header_row(rows)
    if header_idx is None:
        logger.warning("Sheet %s: 'Card #' header not found — skipping", set_code)
        return 0, 0, []

    header = rows[header_idx]
    inv_columns = identify_inventory_columns(header)
    if not inv_columns:
        logger.warning("Sheet %s: no inventory columns in header — skipping", set_code)
        return 0, 0, []

    upserted = 0
    skipped = 0
    failures: list[dict] = []

    for row in rows[header_idx + 1 :]:
        raw_num = row[0]
        if raw_num is None:
            continue

        card_number = parse_card_number(str(raw_num).strip())

        for col_idx, flags in inv_columns.items():
            if col_idx >= len(row):
                continue
            qty_raw = row[col_idx]
            if not is_valid_quantity(qty_raw):
                skipped += 1
                continue

            quantity = int(qty_raw)
            card = _lookup_card(db, set_id, card_number, flags)

            if card is None:
                failures.append(
                    {
                        "set_code": set_code,
                        "card_number": card_number,
                        "column": str(header[col_idx]),
                        "quantity": quantity,
                        "flags": flags,
                    }
                )
                logger.warning(
                    "Lookup failure: set=%s card=%s col=%r",
                    set_code,
                    card_number,
                    header[col_idx],
                )
                continue

            _insert_inventory(db, card.id, quantity)
            upserted += 1

    return upserted, skipped, failures


def find_header_row(rows: list[tuple]) -> int | None:
    """Return the 0-based index of the row whose first cell is 'Card #'."""
    for i, row in enumerate(rows):
        if row[0] is not None and str(row[0]).strip() == "Card #":
            return i
    return None


def identify_inventory_columns(header_row: tuple) -> dict[int, dict[str, bool]]:
    """Return {col_index: flags_dict} for each inventory column in the header.

    Non-inventory columns (Playset, HS-Playset, Rarity, Card Name, etc.) are ignored.
    Matching is case-insensitive to handle variations like 'PRESTIGE' (JTL).
    Only the first occurrence of each column name is used; trailing duplicate
    headers (e.g. summary columns in LOF) are ignored.
    """
    result: dict[int, dict[str, bool]] = {}
    seen: set[str] = set()
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key in VARIANT_COLUMN_FLAGS and key not in seen:
            result[i] = VARIANT_COLUMN_FLAGS[key]
            seen.add(key)
    return result


def is_valid_quantity(value) -> bool:
    """True if value is a positive integer quantity to import.

    Rejects None, strings ('?' playset formula cells), zero, and negatives.
    Accepts integers and floats (openpyxl may return 1.0 for integer cells).
    """
    if value is None or isinstance(value, str):
        return False
    try:
        return int(value) > 0
    except (TypeError, ValueError):
        return False


def _lookup_card(
    db: Session, set_id: int, base_card_number: str, flags: dict[str, bool]
) -> Card | None:
    """Look up a Card by set, base card number, and all five variant flags.

    For JTL/LOF/SEC/LAW (has_unique_variant_numbers=True): base_card_number is the
    standard card's number shared across all variants — lookup is always deterministic.

    For SOR/SHD/TWI (has_unique_variant_numbers=False): Standard and Foil share the
    same base_card_number as the Excel row's card number and resolve correctly.
    Hyperspace and OP variants have distinct card numbers not linked to the base card
    number during F3 ingestion — those lookups fail, are logged, and import continues.

    Promo fallback: when a Promo lookup (is_organized_play=True, is_hyperspace=False)
    finds nothing, retry with is_hyperspace=True. Some sets have OP variants that are
    Hyperspace-only (e.g. SHD Showcase cards), so the Promo column is the correct place
    to track them even though the DB record also carries is_hyperspace=True.
    """
    card = (
        db.query(Card)
        .filter(
            Card.set_id == set_id,
            Card.base_card_number == base_card_number,
            Card.is_foil == flags["is_foil"],
            Card.is_hyperspace == flags["is_hyperspace"],
            Card.is_prestige == flags["is_prestige"],
            Card.is_showcase == flags["is_showcase"],
            Card.is_organized_play == flags["is_organized_play"],
        )
        .one_or_none()
    )

    if card is None and flags["is_organized_play"] and not flags["is_hyperspace"]:
        card = (
            db.query(Card)
            .filter(
                Card.set_id == set_id,
                Card.base_card_number == base_card_number,
                Card.is_foil == flags["is_foil"],
                Card.is_hyperspace == True,
                Card.is_prestige == flags["is_prestige"],
                Card.is_showcase == flags["is_showcase"],
                Card.is_organized_play == True,
            )
            .one_or_none()
        )

    if card is None and flags["is_hyperspace"] and not flags["is_organized_play"]:
        card = (
            db.query(Card)
            .filter(
                Card.set_id == set_id,
                Card.base_card_number == base_card_number,
                Card.is_foil == flags["is_foil"],
                Card.is_hyperspace == True,
                Card.is_prestige == flags["is_prestige"],
                Card.is_showcase == flags["is_showcase"],
                Card.is_organized_play == True,
            )
            .one_or_none()
        )

    return card


def _insert_inventory(db: Session, card_id: int, quantity: int) -> None:
    """Insert an inventory record. Skips if one already exists (DO NOTHING).

    DO NOTHING rather than DO UPDATE: once the UI is live, the database is
    the source of truth for quantities. Overwriting on re-run would silently
    discard UI-managed changes with stale Excel values.
    """
    stmt = pg_insert(Inventory.__table__).values(card_id=card_id, quantity=quantity)
    stmt = stmt.on_conflict_do_nothing(index_elements=["card_id"])
    db.execute(stmt)
